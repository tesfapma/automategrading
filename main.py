# this is an attempt to automate the grading of exams taken for data collectors (Resident Enumerators)


# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

## we are starting to automate the grading system for the RE training exams

# importing required libraries for automating the auto quizzes
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import fonts, Color, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, reference
from openpyxl.utils import FORMULAE
import string
import re

# loading the autoquizzer workbook; workbook1
quiz_excel = pd.read_csv(r'C:\Users\etale\Downloads\CS2020_Mekelle_RE_Final_Exam_V5_results.csv')
# this code is saving the csv file into excel format,i.e., xlsx format
quiz_excel.to_excel(r'C:\Users\etale\Downloads\CS2020_Mekelle_RE_Final_Exam_V5_results.xlsx', index=None, header=True)
# Scratch work, instead of loading the auto quiz and the ODK excel files simultaneously to set a condition for the autoquizzer.
QDK_quiz = load_workbook(r'C:\Users\etale\Desktop\Python test files\PMAET-RE-CS_Training-Final-Mekelle site v4-2020.10.16.xlsx')
QDK_quiz['survey'].insert_cols(3)
for i in range(2, QDK_quiz['survey'].max_row+2):
    if (QDK_quiz['survey'][f'A{i}'].value == f'select_one true_false') or (QDK_quiz['survey'][f'A{i}'].value == f'select_one yes_no') \
            or (QDK_quiz['survey'][f'A{i}'].value == f'decimal') or (QDK_quiz['survey'][f'A{i}'].value == f'integer'):
        QDK_quiz['survey'][f'C{i}'] = QDK_quiz['survey'][f'A{i}'].value + QDK_quiz['survey'][f'B{i}'].value
        # f'=CONCATENATE(A{i},B{i})'
    else:
        QDK_quiz['survey'][f'C{i}'] = QDK_quiz['survey'][f'A{i}'].value
    QDK_quiz.save(r'C:\Users\etale\Desktop\Python test files\PMAET-RE-CS_Training-Final-Mekelle site v4-2020.10.16.xlsx')

Listofpaths = [r'C:\Users\etale\Downloads\CS2020_Mekelle_RE_Final_Exam_V5_results.xlsx',
               r'C:\Users\etale\Desktop\Python test files\PMAET-RE-CS_Training-Final-Mekelle site v4-2020.10.16.xlsx']
#
workbooks = [
    load_workbook(path, data_only=True)
    for path in Listofpaths
]

print(workbooks[1]['survey']['C28'].value)
# wb: Workbook = load_workbook(r'C:\Users\etale\Downloads\CS2020_Mekelle_RE_Final_Exam_V5_results.xlsx')
wb = workbooks[0]
wb1 = workbooks[1]
sheet = wb['Sheet1']
sheet2 = wb1['survey']
sheet3 = wb1['choices']
# sheet = wb.['Sh/eet1']
# cell references (original spreadsheet)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row - 1
# creating a new variable for the question types by concatenating the question type and the question number


# these are codes used to get the list of alphabets for automation process
alphabet = list(string.ascii_lowercase)
excel_alphabet = alphabet[0:max_column]
print(excel_alphabet)

column_list = [cell.column for cell in sheet[1]]
print(column_list)
column_letter = []
for column in column_list:
    column_letter.append(get_column_letter(column))


# getting the cell coordinate where the score of question number "Q1" is located
def find_specific_cell():
    for row in range(1, max_row + 1):
        for column in column_letter:
            cell_name = "{}{}".format(column, row)
            if sheet[cell_name].value == "Q1_score":
                # print("cell position {} has value {}".format(cell_name, sheet[cell_name].value))
                return cell_name


# getting the cell coordinate where the question number "Q1" is located
def find_q1_cell():
    for row in range(1, max_row + 1):
        for column in column_letter:
            cell_name = "{}{}".format(column, row)
            if sheet[cell_name].value == "Q1":
                # print("cell position {} has value {}".format(cell_name, sheet[cell_name].value))
                return cell_name


# loading ODK xlsx file for the quiz, to test the question type

# wb: Workbook = load_workbook(r'C:\Users\etale\Downloads\PMAET-RE-CS_Training-Final-Mekelle site v4-2020.10.16.xlsx')


# Identify the total number of questions
# getting the total number of questions in the ODK format
count = 0
for row_cells in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        cell_coordinate = cell.coordinate
        for match in re.finditer(r'^Q+\d+$', sheet[cell_coordinate].value):
            count += 1

total_question = int(count)
# (f'A{wb1.active.min_row}': f'A{wb1.active.max_row}')
for i in range(1, total_question + 1):
    for row in sheet2.iter_rows(min_col=3, min_row=1, max_col=3, max_row=sheet2.max_row):
        for cell in row:
    # for cell in sheet2['C']:
    #     # print(cell.value)
            if (cell.value == f'select_one Q{i}') or (cell.value == f'select_multiple Q{i}'):
                wb.create_sheet(f'Q{i}')
                wb[f'Q{i}']['A1'] = "Question"
                wb[f'Q{i}']['A2'] = "Average score"
                wb[f'Q{i}']['A3'] = "Answer"
                option_choices = int(0)
                for row in sheet3.iter_rows(min_col=1, min_row=1, max_col=1, max_row=sheet3.max_row):
                    for cell in row:
                        if cell.value == f'Q{i}':
                            option_choices += 1
                for choice in range(4, option_choices + 4):
                    wb[f'Q{i}'][f'A{choice}'] = alphabet[int(f'{choice - 4}')]
                    # wb[f'Q{i}']['A5'] = "b"
                    # wb[f'Q{i}']['A6'] = "c"
                    # wb[f'Q{i}']['A7'] = "d"
                wb[f'Q{i}']['B1'] = f"{i}"
                wb[f'Q{i}']['B2'] = f"=AVERAGE(OFFSET(INDIRECT(\"'Sheet1'!{find_specific_cell()}\"),1,B1-1,{max_row},1))"
                wb[f'Q{i}']['B2'].number_format = numbers.FORMAT_PERCENTAGE_00
                wb[f'Q{i}'].column_dimensions['A'].width = 15
                wb[f'Q{i}']['B3'] = '#'
                for choice in range(4, option_choices + 4):
                    wb[f'Q{i}'][f'B{choice}'] = f"=SUM(ISNUMBER(SEARCH(A{choice}, OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1)))+0)"
                # wb[f'Q{i}']['B5'] = f"=SUM(ISNUMBER(SEARCH(A5, OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1)))+0)"
                # wb[f'Q{i}']['B6'] = f"=SUM(ISNUMBER(SEARCH(A6, OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1)))+0)"
                # wb[f'Q{i}']['B7'] = f"=SUM(ISNUMBER(SEARCH(A7, OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1)))+0)"
                Range = wb[f'Q{i}'][f'B4:B{option_choices + 4}']
                for cell in Range:
                    for x in cell:
                        wb[f'Q{i}'].formula_attributes[x.coordinate] = {'t': 'array','ref': f"{x.coordinate}:{x.coordinate}"}
                wb[f'Q{i}']['C1'] = "Chart Title"
                wb[f'Q{i}']['C3'] = "Percent"
                for choice in range(4, option_choices + 4):
                    wb[f'Q{i}'][f'C{choice}'] = f"=B{choice}/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1-1,{max_row},1))"
                # wb[f'Q{i}']['C5'] = f"=B5/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1-1,{max_row},1))"
                # wb[f'Q{i}']['C6'] = f"=B6/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1-1,{max_row},1))"
                # wb[f'Q{i}']['C7'] = f"=B7/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1-1,{max_row},1))"
                wb[f'Q{i}']['D1'] = '="Q"&B1&": "&TEXT(B2,"0.0%")'
                Range2 = wb[f'Q{i}'][f'C4:C{option_choices + 4}']
                for cell in Range2:
                    for x in cell:
                        wb[f'Q{i}'][x.coordinate].number_format = numbers.FORMAT_PERCENTAGE_00
            elif (cell.value == f'select_one true_falseQ{i}'):
                wb.create_sheet(f'Q{i}')
                wb[f'Q{i}']['A1'] = "Question"
                wb[f'Q{i}']['A2'] = "Average score"
                wb[f'Q{i}']['A3'] = "Answer"
                wb[f'Q{i}']['A4'] = "True"
                wb[f'Q{i}']['A5'] = "False"
                wb[f'Q{i}']['B1'] = f"{i}"
                wb[f'Q{i}']['B2'] = f"=AVERAGE(OFFSET(INDIRECT(\"'Sheet1'!{find_specific_cell()}\"),1,$B$1-1,{max_row},1))"
                wb[f'Q{i}']['B2'].number_format = numbers.FORMAT_PERCENTAGE_00
                wb[f'Q{i}'].column_dimensions['A'].width = 15
                wb[f'Q{i}']['B3'] = '#'
                wb[f'Q{i}']['B4'] = f"=COUNTIF(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,B1 - 1,{max_row},1),TRUE)"
                wb[f'Q{i}']['B5'] = f"=COUNTIF(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,B1 - 1,{max_row},1),FALSE)"
                wb[f'Q{i}']['C1'] = "Chart Title"
                wb[f'Q{i}']['C3'] = "Percent"
                wb[f'Q{i}']['C4'] = f"=B4/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1))"
                wb[f'Q{i}']['C5'] = f"=B5/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1))"
                wb[f'Q{i}']['D1'] = '="Q"&B1&": "&TEXT(B2,"0.0%")'
                Range3 = wb[f'Q{i}'][f'C4:C5']
                for cell in Range3:
                    for x in cell:
                        wb[f'Q{i}'][x.coordinate].number_format = numbers.FORMAT_PERCENTAGE_00
            elif (cell.value == f'select_one yes_noQ{i}'):
                wb.create_sheet(f'Q{i}')
                wb[f'Q{i}']['A1'] = "Question"
                wb[f'Q{i}']['A2'] = "Average score"
                wb[f'Q{i}']['A3'] = "Answer"
                wb[f'Q{i}']['A4'] = "yes"
                wb[f'Q{i}']['A5'] = "no"
                wb[f'Q{i}']['B1'] = f"{i}"
                wb[f'Q{i}']['B2'] = f"=AVERAGE(OFFSET(INDIRECT(\"'Sheet1'!{find_specific_cell()}\"),1,$B$1-1,{max_row},1))"
                wb[f'Q{i}']['B2'].number_format = numbers.FORMAT_PERCENTAGE_00
                wb[f'Q{i}'].column_dimensions['A'].width = 15
                wb[f'Q{i}']['B3'] = '#'
                wb[f'Q{i}']['B4'] = f"=COUNTIF(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1),yes)"
                wb[f'Q{i}']['B5'] = f"=COUNTIF(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1),no)"
                wb[f'Q{i}']['C1'] = "Chart Title"
                wb[f'Q{i}']['C3'] = "Percent"
                wb[f'Q{i}']['C4'] = f"=B4/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1))"
                wb[f'Q{i}']['C5'] = f"=B5/COUNTA(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1 - 1,{max_row},1))"
                wb[f'Q{i}']['D1'] = '="Q"&B1&": "&TEXT(B2,"0.0%")'
                Range3 = wb[f'Q{i}'][f'C4:C5']
                for cell in Range3:
                    for x in cell:
                        wb[f'Q{i}'][x.coordinate].number_format = numbers.FORMAT_PERCENTAGE_00
            elif (cell.value == f'decimalQ{i}') or (cell.value == f'integerQ{i}'):
                wb.create_sheet(f'Q{i}')
                wb[f'Q{i}']['A1'] = "Question"
                wb[f'Q{i}']['A2'] = "Average score"
                wb[f'Q{i}']['A3'] = "Answer"
                wb[f'Q{i}']['A4'] = "True"
                wb[f'Q{i}']['A5'] = "False"
                wb[f'Q{i}']['B1'] = f"{i}"
                wb[f'Q{i}']['B2'] = f"=AVERAGE(OFFSET(INDIRECT(\"'Sheet1'!{find_specific_cell()}\"),1,$B$1-1,{max_row},1))"
                wb[f'Q{i}']['B2'].number_format = numbers.FORMAT_PERCENTAGE_00
                wb[f'Q{i}'].column_dimensions['A'].width = 15
                wb[f'Q{i}']['B3'] = '#'
                wb[f'Q{i}']['D1'] = f"=MIN(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1-1,{max_row},1))"
                wb[f'Q{i}']['D2'] = f"=MAX(OFFSET(INDIRECT(\"'Sheet1'!{find_q1_cell()}\"),1,$B$1-1,{max_row},1))"
                wb[f'Q{i}']['D1'].number_format = numbers.FORMAT_NUMBER
                wb[f'Q{i}']['D2'].number_format = numbers.FORMAT_NUMBER
                print(wb[f'Q{i}']['D2'].value)
                # for choice in range(wb[f'Q{i}']['D1'].value, wb[f'Q{i}']['D2'].value):
                #     wb[f'Q{i}'][f'A{choice}'] = f'{choice}'



            else:
                print("this is not an ODK question: no need to run autouiz ")
wb.save(r'C:\Users\etale\Downloads\CS2020_Mekelle_RE_Final_Exam_V5_results.xlsx')
